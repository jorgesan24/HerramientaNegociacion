import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

BASE_DIR = Path(__file__).resolve().parent
RUTA_PLANTILLA = BASE_DIR / "data" / "plantilla_negociacion.xlsx"

MAPEO_EXPORTACION = {
    1: "CODIGO",
    2: "DESCRIPCION",
    3: "VALOR OFERTADO",
    4: "REGULACION",
    5: "NT",
    6: "REFERENCIA",
    7: "PM",
    8: "VALOR OBJETIVO",
    9: "VALIDACION",
    10: "ESTADO REGISTRO",  # antes INVIMA
    11: "DUPLICADO",
    12: "ORIGEN",           # antes ORIGEN OBJETIVO
    13: "DESVIACION"
}

CAMPOS_ENCABEZADO = {
    # Datos proveedor
    "C8": "identificacion_representante",
    "F8": "nombre_representante",
    "C9": "nit",
    "F9": "proveedor",
    "C10": "ciudad",
    "F10": "sucursal",
    "L10": "codigo_sucursal",
    # Datos contrato
    "J2": "fecha_inicio",
    "E5": "plan_otro",
    "I5": "forma_contratacion"
}

TIPO_ATENCION = {
    "AMBULATORIA": "D3",
    "DOMICILIARIA": "F3",
    "HOSPITALIZACIÓN": "I3",
    "HOSPITALIZACION": "I3",
    "URGENCIAS": "L3"
}

PLAN_CELDAS = {
    "POS CONTRIBUTIVO": "D4",
    "POS SUBSIDIADO": "F4",
    "PLAN EMPRESARIAL": "I4",
    "PLAN PREMIUM": "L4",
}

def marcar_tipo_atencion(ws, encabezado):
    tipos = encabezado.get("tipo_atencion", [])
    if isinstance(tipos, str):
        tipos = [tipos]
    for tipo in tipos:
        tipo = tipo.upper()
        if tipo in TIPO_ATENCION:
            ws[TIPO_ATENCION[tipo]] = "X"

def marcar_plan(ws, encabezado):
    planes = encabezado.get("plan", [])
    if isinstance(planes, str):
        planes = [planes]
    for plan in planes:
        plan = plan.upper()
        if plan in PLAN_CELDAS:
            ws[PLAN_CELDAS[plan]] = "X"
            
    if "OTRO" in [p.upper() for p in planes]:
        ws["D5"] = "X"
        ws["E5"] = encabezado.get("plan_otro", "")

def escribir_encabezado(ws, encabezado):
    for celda, campo in CAMPOS_ENCABEZADO.items():
        ws[celda] = encabezado.get(campo, "")
    marcar_tipo_atencion(ws, encabezado)
    marcar_plan(ws, encabezado)

def generar_excel_negociacion(datos, encabezado, ruta_salida):
    # 1. Asegurar directorios de salida
    carpeta_destino = Path(os.path.dirname(ruta_salida))
    carpeta_destino.mkdir(parents=True, exist_ok=True)
    
    # 2. Cargar la plantilla base solo para modificar las celdas del encabezado
    wb = load_workbook(RUTA_PLANTILLA)
    ws = wb.active
    
    # Escribe los datos fijos del encabezado
    escribir_encabezado(ws, encabezado)
    
    # Guardamos los encabezados de forma temporal antes de inyectar la tabla masiva
    wb.save(ruta_salida)
    wb.close()
    
    # 3. PROCESAMIENTO ULTRA-RÁPIDO CON PANDAS (Vectorización masiva en RAM)
    columnas_ordenadas = [MAPEO_EXPORTACION[i] for i in sorted(MAPEO_EXPORTACION.keys())]
    df_filtrado = datos[columnas_ordenadas].copy()
    
    # Sanitización de datos a alta velocidad con Pandas antes de pasarlo a openpyxl
    for col_name in df_filtrado.columns:
        # Sanitizar columnas de dinero/precios
        if "VALOR" in col_name or "PRECIO" in col_name or "REFERENCIA" in col_name or col_name in ["REGULACION", "NT", "PM"]:
            if df_filtrado[col_name].dtype == object:
                df_filtrado[col_name] = df_filtrado[col_name].astype(str).str.replace('$', '', regex=False)\
                                                             .str.replace('.', '', regex=False)\
                                                             .str.replace(',', '', regex=False).str.strip()
            df_filtrado[col_name] = pd.to_numeric(df_filtrado[col_name], errors="coerce")
            
        # Sanitizar desviaciones o porcentajes
        elif "DESVIACION" in col_name or "PORCENTAJE" in col_name:
            if df_filtrado[col_name].dtype == object:
                df_filtrado[col_name] = df_filtrado[col_name].astype(str).str.replace(',', '.', regex=False).str.strip()
            df_filtrado[col_name] = pd.to_numeric(df_filtrado[col_name], errors="coerce")
            
        # Asegurar cadenas limpias para códigos o textos generales
        else:
            df_filtrado[col_name] = df_filtrado[col_name].fillna("").astype(str).str.strip()

    # Reemplazar valores NaN numéricos por un string vacío compatible con Excel
    df_filtrado = df_filtrado.replace({np.nan: None})

    # Volcar el DataFrame sanitizado a la plantilla de forma atómica
    with pd.ExcelWriter(ruta_salida, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
        df_filtrado.to_excel(
            writer, 
            sheet_name=writer.sheets[ws.title].title, 
            startrow=12,    # Fila 13 en Excel
            startcol=0,     # Columna A
            header=False,   
            index=False     
        )
    
    # 4. APLICACIÓN DE ESTILOS RÁPIDOS EN BLOQUE (Cero procesos lógicos aquí)
    wb_final = load_workbook(ruta_salida)
    ws_final = wb_final.active
    
    aplicar_formatos(ws_final)
    ajustar_columnas(ws_final)
    configurar_impresion(ws_final)
    
    wb_final.save(ruta_salida)
    wb_final.close()
        
    return ruta_salida

def aplicar_formatos(ws):
    """
    Aplica formatos estéticos de forma directa a la cuadrícula. 
    Como los datos ya vienen numéricos y limpios desde Pandas, esta función vuela.
    """
    fuente_datos = Font(name='Calibri', size=9, bold=False, color='000000')
    align_centro = Alignment(horizontal='center', vertical='center')
    align_izquierda = Alignment(horizontal='left', vertical='center')
    align_derecha = Alignment(horizontal='right', vertical='center')
    
    borde_cuadrícula = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    for r_idx in range(13, ws.max_row + 1):
        ws.row_dimensions[r_idx].height = 20  
        
        for col_idx, col_name in MAPEO_EXPORTACION.items():
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.font = fuente_datos
            cell.border = borde_cuadrícula

            # 1. Blindar CUM/NIT
            if col_name in ["CODIGO", "CUM", "NIT"]:
                cell.number_format = '@'  
                cell.alignment = align_centro

            # 2. Formato Moneda
            elif "VALOR" in col_name or "PRECIO" in col_name or "REFERENCIA" in col_name or col_name in ["REGULACION", "NT", "PM"]:
                cell.number_format = '$#,##0'  
                cell.alignment = align_derecha

            # 3. Formato Decimales
            elif "DESVIACION" in col_name or "PORCENTAJE" in col_name:
                cell.number_format = '#,##0.00'  
                cell.alignment = align_derecha

            # 4. Formato Texto General
            else:
                cell.alignment = align_izquierda

def ajustar_columnas(ws):
    """
    Escanea las columnas utilizando los índices rápidos de tupla col[0]
    optimizados para openpyxl moderno.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row >= 12:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
        
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def configurar_impresion(ws):
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE 
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER         